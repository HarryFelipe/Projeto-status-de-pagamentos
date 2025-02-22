'''
1 - Entrar na planilha e extrair o cpf do cliente
2 - Entro no site https://consultcpf-devaprender.netlify.app e uso o cpf da planilha para pesquisar o status do pagamento
3 - Verificar se está em dia ou atrasado
4 - Se estiver em dia, pegar data do pagamento e o método
5 - Caso contrário, colocar status como pendente
6 - Inserir novas informações (nome, valor, cpf, vencimenot, status e caso esteja em dia, data de pagamento, método de pag (cartão ou boleto)) em uma nova planilha
7 - Repetir até chegar o último cliente
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

#1 - Entrar na planilha e extrair o cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app')
sleep(8)

#min_row = é qual a linha mínima que ele vai iniciar
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    #2 - Entro no site https://consultcpf-devaprender.netlify.app e uso o cpf da planilha para pesquisar o status do pagamento

    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    botao_consultar = driver.find_element(By.XPATH, "//*[@id='consultaForm']/button")
    botao_consultar.click()
    sleep(4)
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    if status.text == 'em dia':
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Em dia', data_pagamento_limpo, metodo_pagamento_limpo])
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'Pendente'])